import pandas as pd
import sys
import os
import re
from openpyxl import load_workbook
import shutil

if len(sys.argv) != 3:
    print("Usage: python3 generate_report.py <version> <enforcer>")
    sys.exit(1)

version = sys.argv[1]
enforcer = sys.argv[2]
users = 100

scenarios = {
    "baseline_no_kubearmor": "Without KubeArmor",
    "baseline_kubearmor": "Visibility - None",
    "vis_process": "Visibility - Process",
    "vis_process_file": "Visibility - Process & File",
    "vis_process_network_file": "Visibility - Process, Network & File",
    "vis_network_policy_enforcer": "Visibility - NPE",
    "pol_process": "Policy - Process",
    "pol_process_file": "Policy - Process & File",
    "pol_process_network_file": "Policy - Process, File, Network",
    "pol_network_policy_enforcer": "Policy - NPE",
    "max_load": "Maximum Load"
}

def get_locust_metrics(version, scenario, enforcer):
    file_path = f"results/{version}/{scenario}/{enforcer}/locust_stats.csv"
    if not os.path.exists(file_path):
        return None, None
    try:
        df = pd.read_csv(file_path)
        df['Name'] = df['Name'].astype(str).str.strip()
        agg_row = df[df['Name'] == 'Aggregated']
        if agg_row.empty:
            return None, None
        reqs = round(float(agg_row['Requests/s'].values[0]), 2)
        latency = round(float(agg_row['Average Response Time'].values[0]), 2)
        return reqs, latency
    except:
        return None, None

def get_ka_stats(version, scenario, enforcer):
    file_path = f"results/{version}/{scenario}/{enforcer}/ka_stats.csv"
    if not os.path.exists(file_path):
        return "-", "-"
    try:
        with open(file_path, 'r') as f:
            lines = f.readlines()
        
        data = []
        for line in lines:
            # Strip ANSI and cleanup whitespace
            line = re.sub(r'\x1b\[[0-9;]*[a-zA-Z]', '', line).strip()
            if not line or ',' not in line: continue
            
            parts = line.split(',')
            try:
                cpu = float(parts[0].replace('%', ''))
                
                # Standardize memory units to MiB
                mem_raw = parts[1].split('/')[0].strip() # e.g. "162.4MiB"
                mem_val = float(re.search(r'[\d\.]+', mem_raw).group())
                
                if "GiB" in mem_raw:
                    mem_val *= 1024
                elif "KiB" in mem_raw:
                    mem_val /= 1024
                # if MiB, do nothing
                
                data.append({'CPU': cpu, 'Mem': mem_val})
            except:
                continue
        
        if not data: return "no_data", "no_data"
        df = pd.DataFrame(data)
        return f"{round(df['CPU'].mean(), 2)}%", f"{round(df['Mem'].mean(), 2)} MiB"
    except:
        return "err", "err"

def get_host_metrics(version, scenario, enforcer):
    path = f"results/{version}/{scenario}/{enforcer}"
    vmstat_file = f"{path}/host_vmstat.txt"
    total_mem_file = f"{path}/total_mem.txt"
    
    if not os.path.exists(vmstat_file) or not os.path.exists(total_mem_file):
        return "-", "-", "-"
    try:
        # Get Total RAM from our new file
        with open(total_mem_file, 'r') as f:
            total_ram_mb = float(f.read().strip())
        
        with open(vmstat_file, 'r') as f:
            lines = f.readlines()
        
        rows = []
        for line in lines:
            parts = re.findall(r'\d+', line)
            if len(parts) >= 15: # vmstat typically has 17 columns
                rows.append([float(x) for x in parts])
        
        if len(rows) < 2: return "low_data", "-", "-"
        
        df = pd.DataFrame(rows[1:]) # Skip the first 'boot' row
        
        # vmstat indices (0-based): free=3, buff=4, cache=5, us=12, sy=13
        avg_free_mb  = df[3].mean() / 1024
        avg_buff_mb  = df[4].mean() / 1024
        avg_cache_mb = df[5].mean() / 1024
        
        # Calculate Used: Total - Free - Buffers - Cache
        avg_used_mb = total_ram_mb - avg_free_mb - avg_buff_mb - avg_cache_mb
        
        used_gb = round(avg_used_mb / 1024, 2)
        total_gb = round(total_ram_mb / 1024, 2)
        
        mem_str = f"{used_gb} / {total_gb} GB"
        avg_cpu = f"{round((df[12] + df[13]).mean(), 2)}%"
        avg_io = f"{round((df[8] + df[9]).mean(), 2)} blk/s"
        
        return avg_cpu, mem_str, avg_io
    except:
        return "err", "err", "err"

def generate_excel_report(report_data, version, enforcer):
    template_path = "KubeArmor Performance Benchmarking Report.xlsx" 
    output_path = f"benchmark_report_{version}_{enforcer}.xlsx"
    
    if not os.path.exists(template_path):
        print(f"Error: Template '{template_path}' not found. Cannot generate Excel report.")
        return

    # 1. CREATE the new file by explicitly copying the template
    print(f"Creating new Excel file from template: {output_path}")
    shutil.copy(template_path, output_path)

    # 2. LOAD the newly created file (NOT the template)
    wb = load_workbook(output_path)
    ws = wb.active

    # 3. FILL VALUES
    # Fill the merged title cell (Row 1, Column A)
    ws['A1'] = f"KubeArmor Benchmark Report: {version} ({enforcer.upper()})"

    # Start filling data at Row 3 (since Row 2 has your headers)
    start_row = 3

    for i, row_data in enumerate(report_data):
        current_row = start_row + i
        
        # Mapping data to columns precisely matching your template CSV structure
        ws.cell(row=current_row, column=1, value=row_data["Scenario"])
        ws.cell(row=current_row, column=2, value=row_data["Users"])
        ws.cell(row=current_row, column=3, value=row_data["KA CPU"])
        ws.cell(row=current_row, column=4, value=row_data["KA Memory"])
        ws.cell(row=current_row, column=5, value=row_data["App Throughput (req/s)"])
        ws.cell(row=current_row, column=6, value=row_data["App Latency (ms)"])
        ws.cell(row=current_row, column=7, value=row_data["Throughput Overhead"])
        ws.cell(row=current_row, column=8, value=row_data["Host CPU Util"])
        ws.cell(row=current_row, column=9, value=row_data["Host Memory Usage"])
        ws.cell(row=current_row, column=10, value=row_data["Host Disk I/O"])

    # 4. SAVE the modifications to the new file
    wb.save(output_path)
    print(f"Excel report successfully generated and filled: {output_path}")

# Logic for baseline comparison
baseline_tp, _ = get_locust_metrics(version, "baseline_no_kubearmor", enforcer)

report_data = []

for dir_name, display_name in scenarios.items():
    ka_cpu, ka_mem = get_ka_stats(version, dir_name, enforcer)
    reqs, latency = get_locust_metrics(version, dir_name, enforcer)
    host_cpu, host_mem, host_disk = get_host_metrics(version, dir_name, enforcer)
    
    # Calculate Overhead %
    overhead = "0%"
    if baseline_tp and reqs and dir_name != "baseline_no_kubearmor":
        drop = ((baseline_tp - reqs) / baseline_tp) * 100
        # Format as -X.X% or +X.X% (if performance improved slightly due to noise)
        overhead = f"{drop:.2f}%" if drop >= 0 else f"+{abs(drop):.2f}%"
    elif dir_name == "baseline_no_kubearmor":
        overhead = "N/A"
        ka_cpu, ka_mem = "N/A", "N/A"

    report_data.append({
        "Scenario": display_name,
        "Users": users,
        "KA CPU": ka_cpu,
        "KA Memory": ka_mem,
        "App Throughput (req/s)": reqs if reqs else "-",
        "App Latency (ms)": latency if latency else "-",
        "Throughput Overhead": overhead,
        "Host CPU Util": host_cpu,
        "Host Memory Usage": host_mem,
        "Host Disk I/O": host_disk
    })

# Convert list of dictionaries to a pandas DataFrame
df_report = pd.DataFrame(report_data)

# 1. Generate Markdown Report
md_filename = f"benchmark_report_{version}_{enforcer}.md"
with open(md_filename, "w") as f:
    f.write(f"## KubeArmor Benchmark Report: {version} ({enforcer.upper()})\n\n")
    # Manually write header to match exact requested format
    f.write("| Scenario | Users | KA CPU | KA Memory | App Throughput (req/s) | App Latency (ms) | Host CPU Util | Host Memory Usage | Host Disk I/O |\n")
    f.write("|---|---|---|---|---|---|---|---|---|\n")
    for row in report_data:
        f.write(f"| {row['Scenario']} | {row['Users']} | {row['KA CPU']} | {row['KA Memory']} | {row['App Throughput (req/s)']} | {row['App Latency (ms)']} | {row['Host CPU Util']} | {row['Host Memory Usage']} | {row['Host Disk I/O']} |\n")

print(f"Markdown report saved to {md_filename}")

# 2. Generate Excel Report
generate_excel_report(report_data, version, enforcer)